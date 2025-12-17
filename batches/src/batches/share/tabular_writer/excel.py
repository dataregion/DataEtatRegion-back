from openpyxl import Workbook, load_workbook
import os

from batches.share.tabular_writer.abstract import TabularWriter

class ExcelTabularWriter(TabularWriter):

    def __init__(self, filep: str, username: str | None = None) -> None:
        super().__init__(filep, username)

    def write_header(self, header: list[str]) -> None:
        print(f"ExcelTabularWriter.write_header: {header}")
        
        wb = Workbook()
        ws = wb.active
        
        ws.append(header)
        
        wb.save(self._filep)

    def write_rows(self, rows: list) -> None:
        print(f"ExcelTabularWriter.write_rows: {len(rows)} rows")
        
        if not os.path.exists(self._filep):
            raise FileNotFoundError(f"Le fichier {self._filep} n'existe pas. Appelez write_header d'abord.")
        
        with open(self._filep, 'rb') as f:
            wb = load_workbook(f)
            ws = wb.active
        
            for row in rows:
                ws.append(row)
            
            wb.save(self._filep)

    def close(self) -> None:
        print("ExcelTabularWriter.close")